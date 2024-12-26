import os
import pandas as pd
from openpyxl import load_workbook

# Configurações iniciais
pasta_principal = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios"
caminho_saida_csv = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios\outputs"
caminho_csv_checklist = os.path.join(caminho_saida_csv, "TabelaCheckList.csv")
caminho_csv_erros = os.path.join(caminho_saida_csv, "ErrosCheckList.csv")
os.makedirs(caminho_saida_csv, exist_ok=True)

# Função para ajustar permissões
def ajustar_permissoes(pasta):
    for root, _, files in os.walk(pasta):
        for arquivo in files:
            caminho_arquivo = os.path.join(root, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                    os.chmod(caminho_arquivo, 0o666)
                except Exception as e:
                    print(f"Erro ao ajustar permissões para {arquivo}: {e}")

# Função para registrar erros
def registrar_erros(erros):
    if erros:
        df_erros = pd.DataFrame(erros)
        df_erros.to_csv(caminho_csv_erros, index=False, encoding="utf-8-sig")
        print(f"Relatório de erros gerado em: {caminho_csv_erros}")
    else:
        print("Nenhum erro encontrado.")

# Função para processar CheckLists
def processar_checklist(caminho_arquivo):
    erros = []
    dados = []
    print(f"Processando arquivo: {caminho_arquivo}")
    try:
        workbook = load_workbook(caminho_arquivo, data_only=True)

        # Processa a tabela "2- Check List 01"
        if "2- Check List 01" in workbook.sheetnames:
            print(f"Planilha encontrada: 2- Check List 01")
            planilha = workbook["2- Check List 01"]
            referencias_checklist01 = [
                25, 28, 33, 36, 39, 44, 47, 50, 55, 58, 61, 64, 69, 72, 75, 80, 83, 86, 91
            ]
            for row in referencias_checklist01:
                referencia_nr12 = planilha[f"D{row}"].value
                status = planilha[f"CN{row}"].value

                if referencia_nr12 and "̴" in referencia_nr12 and status == "Ok":
                    print(f"Processando intervalo: {referencia_nr12}")
                    ranges = referencia_nr12.split(" ̴ ")
                    try:
                        inicio, fim = map(str.strip, ranges)
                        inicio_valor = int(inicio.replace(".", "").replace(" ", ""))
                        fim_valor = int(fim.replace(".", "").replace(" ", ""))
                        for ref in range(inicio_valor, fim_valor + 1):
                            referencia_formatada = f"{ref // 100}.{ref % 100}"
                            dados.append({"Planilha": "2- Check List 01", "Referência NR-12": referencia_formatada, "Status": status})
                    except ValueError:
                        erros.append({"Arquivo": caminho_arquivo, "Erro": f"Formato inválido: {referencia_nr12}"})

        # Processa a tabela "3- Check List 02"
        if "3- Check List 02" in workbook.sheetnames:
            print(f"Planilha encontrada: 3- Check List 02")
            planilha = workbook["3- Check List 02"]
            referencias_checklist02 = [25, 28, 31, 36, 41]
            for row in referencias_checklist02:
                referencia_nr12 = planilha[f"D{row}"].value
                status = planilha[f"CN{row}"].value

                if referencia_nr12 and "̴" in referencia_nr12 and status == "Ok":
                    print(f"Processando intervalo: {referencia_nr12}")
                    ranges = referencia_nr12.split(" ̴ ")
                    try:
                        inicio, fim = map(str.strip, ranges)
                        inicio_valor = int(inicio.replace(".", "").replace(" ", ""))
                        fim_valor = int(fim.replace(".", "").replace(" ", ""))
                        for ref in range(inicio_valor, fim_valor + 1):
                            referencia_formatada = f"{ref // 100}.{ref % 100}"
                            dados.append({"Planilha": "3- Check List 02", "Referência NR-12": referencia_formatada, "Status": status})
                    except ValueError:
                        erros.append({"Arquivo": caminho_arquivo, "Erro": f"Formato inválido: {referencia_nr12}"})

    except Exception as e:
        erros.append({"Arquivo": caminho_arquivo, "Erro": str(e)})

    return dados, erros

# Processar os arquivos Excel
ajustar_permissoes(pasta_principal)
erros_gerais = []

for root, _, files in os.walk(pasta_principal):
    for arquivo in files:
        if arquivo.endswith((".xlsx", ".xlsm")):
            caminho_arquivo = os.path.join(root, arquivo)
            try:
                dados, erros = processar_checklist(caminho_arquivo)

                # Salvar os dados no CSV
                if dados:
                    df = pd.DataFrame(dados)
                    if os.path.exists(caminho_csv_checklist):
                        df.to_csv(caminho_csv_checklist, mode="a", header=False, index=False, encoding="utf-8-sig")
                    else:
                        df.to_csv(caminho_csv_checklist, index=False, encoding="utf-8-sig")
                    print(f"Dados salvos no CSV: {caminho_csv_checklist}")

                # Registrar erros
                erros_gerais.extend(erros)
            except Exception as e:
                erros_gerais.append({"Arquivo": arquivo, "Caminho": caminho_arquivo, "Erro": str(e)})

# Gerar relatório geral de erros
registrar_erros(erros_gerais)

print(f"Tabela CheckList gerada em: {caminho_csv_checklist}")
