import os
import pandas as pd
from openpyxl import load_workbook

# Caminhos fornecidos
pasta_principal = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios"  # Pasta principal
caminho_saida_csv = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios\outputs"
os.makedirs(caminho_saida_csv, exist_ok=True)  # Cria a pasta de saída, se não existir

# Nome do CSV de saída
nome_csv = "TabelaMaquina.csv"
nome_erro_csv = "RelatorioErros.csv"
controle_processados = "processados.txt"  # Arquivo de controle para arquivos processados

# Mapeamento das células da planilha "1- Capa"
celulas_capa = {
    "BD35": "AnoFabricacao",
    "BD26": "Capacidade",
    "CN70": "Categoria Risco",
    "BR75": "Categoria SIL",
    "BB58": "F1",
    "BB61": "F2",
    "V13": "Fabricante",
    "BS10": "Local",
    "BS10": "Local_en",
    "BS13": "Modelo",
    "V10": "Nome",
    "V11": "Nome_En",
    "BS16": "NumIndice",
    "V16": "NumSerie",
    "BW26": "OutrasInformacoes",
    "BB67": "P1",
    "BB70": "P2",
    "CN75": "PerformanceLevel",
    "BD32": "Peso",
    "BD29": "Potencia",
    "BB49": "S1",
    "BB52": "S2",
    "BD38": "StatusNr12",
    "BD23": "TagCliente",
}

# Células para contagem na planilha "2- Check List 01"
celulas_checklist = [
    "CN25", "CN28", "CN33", "CN36", "CN39", "CN44", "CN47", "CN50", "CN55",
    "CN58", "CN61", "CN64", "CN69", "CN72", "CN75", "CN80", "CN83", "CN86", "CN91"
]

# Função para ajustar permissões dos arquivos na pasta
def ajustar_permissoes(pasta):
    for root, _, files in os.walk(pasta):
        for arquivo in files:
            caminho_arquivo = os.path.join(root, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                    os.chmod(caminho_arquivo, 0o666)  # Permissão de leitura e escrita
                except Exception as e:
                    print(f"Erro ao ajustar permissões para {arquivo}: {e}")

# Função para processar um arquivo Excel
def processar_excel(caminho_arquivo):
    nome_arquivo = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    workbook = load_workbook(caminho_arquivo, data_only=True)

    # Inicializar o dicionário de dados
    dados = {"Arquivo": nome_arquivo}

    # Processar a planilha "1- Capa"
    if "1- Capa" in workbook.sheetnames:
        planilha_capa = workbook["1- Capa"]
        for celula, nome_coluna in celulas_capa.items():
            valor = planilha_capa[celula].value
            dados[nome_coluna] = valor

    # Processar a planilha "2- Check List 01"
    if "2- Check List 01" in workbook.sheetnames:
        planilha_checklist = workbook["2- Check List 01"]
        soma_na, soma_nc, soma_ok = 0, 0, 0
        for celula in celulas_checklist:
            valor = str(planilha_checklist[celula].value).strip().upper() if planilha_checklist[celula].value else ""
            if valor == "NA":
                soma_na += 1
            elif valor == "NC":
                soma_nc += 1
            elif valor == "OK":
                soma_ok += 1
        dados["SomaNA"], dados["SomaNc"], dados["SomaOk"] = soma_na, soma_nc, soma_ok

    return dados

# Função para ler arquivos processados
def ler_processados():
    if os.path.exists(controle_processados):
        with open(controle_processados, "r") as f:
            return set(f.read().splitlines())
    return set()

# Função para registrar arquivos processados
def registrar_processado(arquivo):
    with open(controle_processados, "a") as f:
        f.write(arquivo + "\n")

# Ajustar permissões dos arquivos na pasta principal e subpastas
ajustar_permissoes(pasta_principal)

# Listas para armazenar os dados
todos_dados = []
erros = []

# Ler arquivos já processados
arquivos_processados = ler_processados()

# Processar todos os arquivos Excel na pasta principal e subpastas
for root, _, files in os.walk(pasta_principal):
    for arquivo in files:
        if arquivo.endswith((".xlsx", ".xlsm")):
            caminho_arquivo = os.path.join(root, arquivo)
            if caminho_arquivo not in arquivos_processados:
                try:
                    dados = processar_excel(caminho_arquivo)
                    if dados:
                        todos_dados.append(dados)
                        registrar_processado(caminho_arquivo)
                        # Salvar progressivamente
                        caminho_csv = os.path.join(caminho_saida_csv, nome_csv)
                        df = pd.DataFrame(todos_dados)
                        df.to_csv(caminho_csv, index=False, encoding="utf-8-sig")
                        print(f"Progresso salvo para: {caminho_arquivo}")
                except PermissionError:
                    erros.append({"Arquivo não processado": arquivo, "Caminho": caminho_arquivo})
                except Exception as e:
                    erros.append({"Arquivo não processado": arquivo, "Caminho": caminho_arquivo, "Erro": str(e)})

# Gerar CSV com o relatório de erros
if erros:
    caminho_erro_csv = os.path.join(caminho_saida_csv, nome_erro_csv)
    df_erros = pd.DataFrame(erros)
    df_erros.to_csv(caminho_erro_csv, index=False, encoding="utf-8-sig")
    print(f"Relatório de erros gerado em: {caminho_erro_csv}")
else:
    print("Nenhum erro registrado.")
