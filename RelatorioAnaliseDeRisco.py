import os
import pandas as pd
from openpyxl import load_workbook

# Configurações iniciais
pasta_principal = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios"
caminho_saida_csv = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios\outputs"
caminho_csv_analise = os.path.join(caminho_saida_csv, "RelatorioAnaliseDeRisco.csv")
caminho_csv_erros_analise = os.path.join(caminho_saida_csv, "ErrosAnaliseDeRisco.csv")
os.makedirs(caminho_saida_csv, exist_ok=True)

# Função para ajustar permissões
def ajustar_permissoes(pasta):
    for root, _, files in os.walk(pasta):
        for arquivo in files:
            caminho_arquivo = os.path.join(root, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                    os.chmod(caminho_arquivo, 0o666)
                except Exception as e:
                    print(f"Erro ao ajustar permissões para {arquivo}: {e}")

# Função para registrar erros
def registrar_erros(erros, caminho_arquivo_erro):
    if erros:
        df_erros = pd.DataFrame(erros)
        df_erros.to_csv(caminho_arquivo_erro, index=False, encoding="utf-8-sig")
        print(f"Relatório de erros gerado em: {caminho_arquivo_erro}")
    else:
        print("Nenhum erro encontrado.")

# Função para processar análise de risco
def processar_analise(caminho_arquivo):
    erros = []
    dados = []
    print(f"Processando arquivo: {caminho_arquivo}")
    try:
        workbook = load_workbook(caminho_arquivo, data_only=True)

        if "1- Capa" in workbook.sheetnames:
            print(f"Planilha encontrada: 1- Capa")
            planilha = workbook["1- Capa"]

            # Definir as células a serem coletadas
            nomes = ["I86", "I88", "I90", "I92", "I94"]
            locais = ["AH86", "AH88", "AH90", "AH92", "AH94"]
            fe = ["BK86", "BK88", "BK90", "BK92", "BK94"]

            for nome_cell, local_cell, fe_cell in zip(nomes, locais, fe):
                nome = planilha[nome_cell].value
                local = planilha[local_cell].value
                fe_valor = planilha[fe_cell].value

                # Adicionar ao relatório apenas se todos os campos estiverem preenchidos
                if nome and local and fe_valor:
                    dados.append({
                        "Arquivo": os.path.basename(caminho_arquivo),
                        "Nome": nome,
                        "Local": local,
                        "FE": fe_valor,
                    })

    except Exception as e:
        erros.append({"Arquivo": caminho_arquivo, "Erro": str(e)})

    return dados, erros

# Processar os arquivos Excel
ajustar_permissoes(pasta_principal)
erros_gerais = []

for root, _, files in os.walk(pasta_principal):
    for arquivo in files:
        if arquivo.endswith((".xlsx", ".xlsm")):
            caminho_arquivo = os.path.join(root, arquivo)
            try:
                dados, erros = processar_analise(caminho_arquivo)

                # Salvar os dados no CSV
                if dados:
                    df = pd.DataFrame(dados)
                    if os.path.exists(caminho_csv_analise):
                        df.to_csv(caminho_csv_analise, mode="a", header=False, index=False, encoding="utf-8-sig")
                    else:
                        df.to_csv(caminho_csv_analise, index=False, encoding="utf-8-sig")
                    print(f"Dados salvos no CSV: {caminho_csv_analise}")

                # Registrar erros
                erros_gerais.extend(erros)
            except Exception as e:
                erros_gerais.append({"Arquivo": arquivo, "Caminho": caminho_arquivo, "Erro": str(e)})

# Gerar relatório geral de erros
registrar_erros(erros_gerais, caminho_csv_erros_analise)

print(f"Relatório de Análise de Risco gerado em: {caminho_csv_analise}")
