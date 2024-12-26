import os
import pandas as pd
from openpyxl import load_workbook

# Configurações iniciais
pasta_principal = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios"
caminho_saida_csv = r"C:\Users\Raphael Medeiros\Desktop\07 - Relatórios\outputs"
caminho_csv_acoes = os.path.join(caminho_saida_csv, "TabelaAcoes.csv")
caminho_csv_erros_acoes = os.path.join(caminho_saida_csv, "ErrosAcoes.csv")
os.makedirs(caminho_saida_csv, exist_ok=True)

# Função para ajustar permissões
def ajustar_permissoes(pasta):
    for root, _, files in os.walk(pasta):
        for arquivo in files:
            caminho_arquivo = os.path.join(root, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                    os.chmod(caminho_arquivo, 0o666)
                except Exception as e:
                    print(f"Erro ao ajustar permissões para {arquivo}: {e}")

# Função para registrar erros
def registrar_erros(erros, caminho_arquivo_erro):
    if erros:
        df_erros = pd.DataFrame(erros)
        df_erros.to_csv(caminho_arquivo_erro, index=False, encoding="utf-8-sig")
        print(f"Relatório de erros gerado em: {caminho_arquivo_erro}")
    else:
        print("Nenhum erro encontrado.")

# Função para processar ações
def processar_acoes(caminho_arquivo):
    erros = []
    dados = []
    print(f"Processando arquivo: {caminho_arquivo}")
    try:
        workbook = load_workbook(caminho_arquivo, data_only=True)

        if "4A- Não Conform." in workbook.sheetnames:
            print(f"Planilha encontrada: 4A- Não Conform.")
            planilha = workbook["4A- Não Conform."]

            # Coleta os dados específicos
            descricao = planilha["D73"].value
            descricao_en = planilha["D75"].value

            # Adiciona ao conjunto de dados
            dados.append({
                "descricao": descricao if descricao else "",
                "descricao_en": descricao_en if descricao_en else "",
                "maquina": "",
                "nc": "",
                "Projeto": "",
                "ref": "",
                "Responsavel": "",
            })

    except Exception as e:
        erros.append({"Arquivo": caminho_arquivo, "Erro": str(e)})

    return dados, erros

# Processar os arquivos Excel
ajustar_permissoes(pasta_principal)
erros_gerais = []

for root, _, files in os.walk(pasta_principal):
    for arquivo in files:
        if arquivo.endswith((".xlsx", ".xlsm")):
            caminho_arquivo = os.path.join(root, arquivo)
            try:
                dados, erros = processar_acoes(caminho_arquivo)

                # Salvar os dados no CSV
                if dados:
                    df = pd.DataFrame(dados)
                    if os.path.exists(caminho_csv_acoes):
                        df.to_csv(caminho_csv_acoes, mode="a", header=False, index=False, encoding="utf-8-sig")
                    else:
                        df.to_csv(caminho_csv_acoes, index=False, encoding="utf-8-sig")
                    print(f"Dados salvos no CSV: {caminho_csv_acoes}")

                # Registrar erros
                erros_gerais.extend(erros)
            except Exception as e:
                erros_gerais.append({"Arquivo": arquivo, "Caminho": caminho_arquivo, "Erro": str(e)})

# Gerar relatório geral de erros
registrar_erros(erros_gerais, caminho_csv_erros_acoes)

print(f"Tabela Ações gerada em: {caminho_csv_acoes}")
